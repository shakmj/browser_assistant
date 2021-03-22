# import library
from tkinter import *
import tkinter as tk
import glob
from functools import partial
from selenium import webdriver
from openpyxl import load_workbook

##
print("welcome to browser assistant... ")

## global configuration variable, to configure screen size.
button_space = 50 # space betwen button
# declare my firefox profile file to this program
#currently not work
#my_profile = webdriver.FirefoxProfile(
#    r"c:\Users\")
#driver = webdriver.Firefox(my_profile)
#Initialization
#initiat web driver from selnieum library
driver = webdriver.Firefox()
print("driver initialized.")
##global configuration variable, to configure screen size.

# functions
# open specific website page.
def forward_page(url_index, url_list):
    # url_list : list contain all url's of category , url_ index : the index for needed site
    print("url index " + str(url_index))
    print("page will opened: " +url_list[url_index])
    driver.get(url_list[url_index])
    return

# create category choice screen.
# when pressing on category button, this function will make transition to show the submenu.
def toscreen(category_choiced):
    global secscreen                             # the second screen we created before.
    secscreen = tk.Tk()                          # i dont know why this here also.
    print("category choiced: "+category_choiced)
    global flag_of_state                         # flag of program state to control the flow of program
    flag_of_state = 3                            # cuurent proses flag
    pages, url = openfile(category_choiced)      # open file contain web page of selected category
    number_of_pages = len(pages)                 # number of webpage to automaticly configure screen .
    if number_of_pages < 10:                     # screen configure proses
        M = number_of_pages * button_space
        N = 200
    else:
        M = number_of_pages * button_space
        N = 400
    category_choicedname = category_choiced + "menu"
    # after collecting all needed data , let's create the submenu
    obj = init_secsc(category_choicedname, M, N, pages, category_choiced, forward_page, url)
    #def init_secsc(screen_name, M, N, choices, label, fun, url):

    return category_choiced
# get xlsx file around the program to create main menu.
def getfilesname():
    # read list of xls in current folder
    xls_list = glob.glob("*.xlsx")
    number_of_choices = len(xls_list)
    # m,n its the screen size .
    if number_of_choices < 10:
        M = number_of_choices * button_space
        N = 200
    else:
        M = number_of_choices * button_space
        N = 400
    return xls_list, number_of_choices, M, N

# create main screen.
def init_mainsc(screen_name, M, N, choices, label, fun):
    button_obj = []
    screen_name = tk.Tk()
    screen_name.wm_title(label)
    # configure geometry
    geo = str(N) + "x" + str(M)
    screen_name.geometry(geo)
    #create button on screen related to submenus.
    for i, obj in enumerate(choices):
        print(i)
        button_obj.append(tk.Button(screen_name, text=obj, command=partial(fun, obj)))
    # configure window grid (not work correctly).
    for i, btn in enumerate(button_obj):
        btn.grid(row=i + 1, column=2)
    return screen_name

#create submenu screen.
def init_secsc(screen_name, M, N, choices, label, fun, url):
    global secscreen # i dont know why this here!!
    button_obj = []
    secscreen.wm_title(label)
    # configure geometry
    geo = str(N) + "x" + str(M)
    secscreen.geometry(geo)
    #create button on screen related to webpage.
    for i, obj in enumerate(choices):
        print(i)
        button_obj.append(tk.Button(secscreen, text=obj, command=partial(fun, i, url)))
    # configure window
    for i, btn in enumerate(button_obj):
        btn.grid(row=i + 1, column=2)
    return secscreen

# load and read the selected xlsx file.
def openfile(file_choiced):
    #load xlsx file.
    workbook = load_workbook(filename=file_choiced)
    #activate it.
    sheet = workbook.active
    #print the title of sheet
    print(sheet.title)
    #get the number of rows
    sheet_row = sheet.max_row
    print(sheet_row)
    pages = []
    url = []
    #add the pages and url's from xlsx to list we created before.
    for i in range(1, sheet_row + 1, 1):
        # print(i)
        cells_page = sheet.cell(row=i, column=1).value
        cells_url = sheet.cell(row=i, column=2).value
        pages.append(cells_page)
        url.append(cells_url)
    print(pages)
    print(url)
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    return pages, url

### main loop
button_obj_mainsc = []
flag_of_state = 1
bast_f = 1
# read the file in cuurent folder
categores, numnerOfChoices, m, n = getfilesname()
# create the main screen
main_sc = init_mainsc("main_sc", m, n, categores, "main categories", toscreen)

while True:
    while flag_of_state == 1: # first pros
        main_sc.update()
        print(str(main_sc.winfo_exists()))
        if str(main_sc.winfo_exists())=="1":
            main_sc.deiconify()
    while flag_of_state == 3:
        try:
            secscreen.update()
            main_sc.withdraw()
        except:
            flag_of_state =1

